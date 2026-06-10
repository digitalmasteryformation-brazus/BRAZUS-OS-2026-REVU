#!/usr/bin/env python3
"""
BRASUS - Google Workspace Files Generator v2.1
Digital Mastery Agency (DMA) - Abidjan, Côte d'Ivoire
Génère les fichiers Excel (.xlsx) importables dans Google Sheets
avec protection réelle des cellules en lecture seule.
"""

import os
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter

# ── PALETTE ─────────────────────────────────────────────────────────────
HDR_BLUE     = "0D47A1"
MID_BLUE     = "1565C0"
SUBHDR_BLUE  = "283593"
ACCENT_BLUE  = "1976D2"
LIGHT_BLUE   = "E3F2FD"
DARK_GREY    = "37474F"
MID_GREY     = "B0BEC5"
LIGHT_GREY   = "ECEFF1"
LOCKED_BG    = "CFD8DC"   # fond gris = cellule verrouillée
INPUT_YLW    = "FFFDE7"   # fond jaune = cellule saisissable
WHITE        = "FFFFFF"
GOLD         = "F9A825"
LIGHT_GOLD   = "FFF9C4"
GREEN_DARK   = "1B5E20"
GREEN_LIGHT  = "E8F5E9"
GREEN_MID    = "2E7D32"
RED_DARK     = "B71C1C"
RED_LIGHT    = "FFEBEE"
ORANGE_DARK  = "E65100"
ORANGE_LIGHT = "FFF3E0"
PURPLE_MID   = "6A1B9A"
PURPLE_LIGHT = "F3E5F5"

# Mot de passe admin pour déverrouiller les feuilles
# ⚠️ À CHANGER avant déploiement client
ADMIN_PWD = "BRASUS_ADMIN_2026"

# ── HELPERS ─────────────────────────────────────────────────────────────
def fill(c):
    return PatternFill(start_color=c, end_color=c, fill_type="solid")

def font(c="000000", sz=10, bold=False, italic=False):
    return Font(color=c, size=sz, bold=bold, italic=italic, name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def lborder(color="000000"):
    return Border(left=Side(style="medium", color=color),
                  right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

def cell(ws, r, c, val="", f=None, ft=None, al=None, bd=None, locked=True):
    """Crée une cellule. locked=True → verrouillée, locked=False → saisissable."""
    cx = ws.cell(row=r, column=c, value=val)
    if f:  cx.fill = f
    if ft: cx.font = ft
    if al: cx.alignment = al
    if bd: cx.border = bd
    cx.protection = Protection(locked=locked)
    return cx

def merge(ws, r1, c1, r2, c2, val="", f=None, ft=None, al=None, bd=None, locked=True):
    """Fusionne et remplit. locked=True → verrouillée, locked=False → saisissable."""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cx = ws.cell(row=r1, column=c1, value=val)
    if f:  cx.fill = f
    if ft: cx.font = ft
    if al: cx.alignment = al
    if bd: cx.border = bd
    cx.protection = Protection(locked=locked)
    return cx

def enable_protection(ws):
    """Active la protection de la feuille — seules les cellules locked=False restent éditables."""
    ws.protection.sheet = True
    ws.protection.password = ADMIN_PWD
    ws.protection.selectLockedCells = True    # lecture autorisée
    ws.protection.selectUnlockedCells = True  # saisie autorisée sur cellules débloquées
    ws.protection.formatCells = True          # empêche modification de format
    ws.protection.formatColumns = True
    ws.protection.formatRows = True
    ws.protection.insertColumns = True
    ws.protection.insertRows = True
    ws.protection.deleteColumns = True
    ws.protection.deleteRows = True

def hrow(ws, r, h):
    ws.row_dimensions[r].height = h

def wcol(ws, col_letter, w):
    ws.column_dimensions[col_letter].width = w

# ── ÉLÉMENTS STRUCTURELS (tous locked=True) ─────────────────────────────
def banner(ws, r, cols, title, bg=HDR_BLUE, sz=16):
    hrow(ws, r, 42)
    merge(ws, r, 1, r, cols, title,
          f=fill(bg), ft=font(WHITE, sz, bold=True), al=align("center","center"),
          locked=True)

def subbar_split(ws, r, c_mid, total, left, right="", bg=SUBHDR_BLUE):
    hrow(ws, r, 24)
    merge(ws, r, 1, r, c_mid, left,
          f=fill(bg), ft=font(WHITE, 10), al=align("left","center"), locked=True)
    if right:
        merge(ws, r, c_mid+1, r, total, right,
              f=fill(bg), ft=font(LIGHT_GOLD, 9, italic=True), al=align("right","center"), locked=True)

def section_header(ws, r, cols, title, bg=DARK_GREY):
    hrow(ws, r, 28)
    merge(ws, r, 1, r, cols, f"  {title}",
          f=fill(bg), ft=font(WHITE, 11, bold=True), al=align("left","center"), locked=True)

def spacer(ws, r, h=10):
    hrow(ws, r, h)

def footer_lock(ws, r, cols, msg="🔒 Feuille en lecture seule — Mise à jour automatique par BRASUS"):
    hrow(ws, r, 22)
    merge(ws, r, 1, r, cols, msg,
          f=fill(LOCKED_BG), ft=font(DARK_GREY, 9, italic=True), al=align("center"), locked=True)

# ════════════════════════════════════════════════════════════════════════
# SHEET 1 — MISSIONS DU JOUR
# ════════════════════════════════════════════════════════════════════════
def sheet_missions(ws, dept):
    wcol(ws, "A", 5); wcol(ws, "B", 46); wcol(ws, "C", 17)
    wcol(ws, "D", 29); wcol(ws, "E", 31); wcol(ws, "F", 20); wcol(ws, "G", 36)

    banner(ws, 1, 7, f"🎯  MISSIONS DU JOUR — {dept.upper()}")
    hrow(ws, 2, 24)
    merge(ws, 2, 1, 2, 4, "DATE : {{DATE_DU_JOUR}}     RESPONSABLE : {{PRÉNOM}}",
          f=fill(SUBHDR_BLUE), ft=font(WHITE, 10), al=align("left","center"), locked=True)
    merge(ws, 2, 5, 2, 7, "⚠️ Missions générées automatiquement par BRASUS — Non modifiables",
          f=fill(SUBHDR_BLUE), ft=font(LIGHT_GOLD, 9, italic=True), al=align("right","center"), locked=True)
    spacer(ws, 3, 8)

    hrow(ws, 4, 30)
    for c, h in enumerate(["#","MISSION","PRIORITÉ","OBJECTIF LIÉ","RÉSULTAT ATTENDU","STATUT","COMMENTAIRE"], 1):
        cell(ws, 4, c, h, f=fill(DARK_GREY), ft=font(WHITE, 10, bold=True),
             al=align("center","center",True), bd=border(), locked=True)

    prio_cfg = [
        ("🔴 Critique", RED_LIGHT),
        ("🔴 Critique", RED_LIGHT),
        ("🟡 Haute",    LIGHT_GOLD),
        ("🟡 Haute",    LIGHT_GOLD),
        ("🟢 Standard", GREEN_LIGHT),
        ("🟢 Standard", GREEN_LIGHT),
        ("🟢 Standard", GREEN_LIGHT),
        ("🟢 Standard", GREEN_LIGHT),
    ]
    for i, (prio, pbg) in enumerate(prio_cfg, 1):
        r = 4 + i; hrow(ws, r, 36)
        # ── Colonnes A–E : VERROUILLÉES (générées par BRASUS) ──
        cell(ws, r, 1, i,    f=fill(LOCKED_BG), ft=font(sz=11, bold=True), al=align("center"), bd=border(), locked=True)
        cell(ws, r, 2, "[Mission générée par BRASUS à 07:00]",
             f=fill(LOCKED_BG), ft=font(italic=True,sz=10), al=align("left","center",True), bd=border(), locked=True)
        cell(ws, r, 3, prio, f=fill(LOCKED_BG), ft=font(bold=True,sz=10), al=align("center"), bd=border(), locked=True)
        cell(ws, r, 4, "[Objectif stratégique lié]",
             f=fill(LOCKED_BG), ft=font(italic=True,sz=9), al=align("left","center",True), bd=border(), locked=True)
        cell(ws, r, 5, "[Résultat mesurable attendu]",
             f=fill(LOCKED_BG), ft=font(italic=True,sz=9), al=align("left","center",True), bd=border(), locked=True)
        # ── Colonnes F–G : SAISISSABLES par le responsable ──
        cell(ws, r, 6, "☐ Non commencé",
             f=fill(INPUT_YLW), ft=font(sz=10), al=align("center"), bd=border(), locked=False)
        cell(ws, r, 7, "",
             f=fill(INPUT_YLW), ft=font(sz=10), al=align("left","center",True), bd=border(), locked=False)

    hrow(ws, 13, 8)
    hrow(ws, 14, 22)
    merge(ws, 14, 1, 14, 7,
          "STATUTS : ☐ Non commencé  |  🔄 En cours  |  ✅ Réalisé  |  ❌ Non réalisé",
          f=fill(LIGHT_GREY), ft=font(DARK_GREY, 9, italic=True), al=align("center"), locked=True)
    hrow(ws, 15, 22)
    merge(ws, 15, 1, 15, 7,
          "🔒 Colonnes A–E verrouillées (BRASUS) — Colonnes STATUT et COMMENTAIRE saisissables (fond jaune)",
          f=fill(ORANGE_LIGHT), ft=font(ORANGE_DARK, 9, italic=True), al=align("center"), locked=True)
    hrow(ws, 16, 22)
    merge(ws, 16, 1, 16, 7,
          "⏰ Verrouillage automatique à 18h00 (heure d'Abidjan) — Toutes les cellules verrouillées après clôture",
          f=fill(RED_LIGHT), ft=font(RED_DARK, 9, italic=True), al=align("center"), locked=True)

    enable_protection(ws)

# ════════════════════════════════════════════════════════════════════════
# SHEET 2 — EXÉCUTION JOURNALIÈRE
# ════════════════════════════════════════════════════════════════════════
def sheet_execution(ws, dept):
    wcol(ws, "A", 30); wcol(ws, "B", 65); wcol(ws, "C", 22)

    banner(ws, 1, 3, f"📋  EXÉCUTION JOURNALIÈRE — {dept.upper()}")
    hrow(ws, 2, 24)
    merge(ws, 2, 1, 2, 2, "DATE : {{DATE_DU_JOUR}}",
          f=fill(SUBHDR_BLUE), ft=font(WHITE, 11), al=align("left","center"), locked=True)
    cell(ws, 2, 3, f"DÉPARTEMENT : {dept}",
         f=fill(SUBHDR_BLUE), ft=font(WHITE, 11), al=align("right","center"), locked=True)
    spacer(ws, 3, 10)

    def section_block(r, title, color, note=None):
        hrow(ws, r, 28)
        merge(ws, r, 1, r, 3, f"  {title}",
              f=fill(color), ft=font(WHITE, 12, bold=True), al=align("left","center"), locked=True)
        if note:
            hrow(ws, r+1, 18)
            merge(ws, r+1, 1, r+1, 3, note,
                  f=fill(LIGHT_GREY), ft=font(DARK_GREY, 9, italic=True), al=align("left","center"), locked=True)
            return r+2
        return r+1

    # Section A — Actions Réalisées (SAISISSABLE)
    r_a = section_block(4, "SECTION A — Actions Réalisées  ★ OBLIGATOIRE", MID_BLUE,
                         "Décrivez toutes les actions réalisées aujourd'hui (minimum 3 lignes)")
    for r in range(r_a, r_a+5):
        hrow(ws, r, 33)
        merge(ws, r, 1, r, 3, "", f=fill(INPUT_YLW), ft=font(sz=10),
              al=align("left","center",True), locked=False)
        ws.cell(r, 1).border = border()

    spacer(ws, r_a+5, 10)

    # Section B — Résultats par mission (SAISISSABLE colonnes B et C)
    r_b = section_block(r_a+6, "SECTION B — Résultats Obtenus (par Mission)", MID_BLUE)
    hrow(ws, r_b, 24)
    for c, h in enumerate(["MISSION", "RÉSULTAT OBTENU / COMMENTAIRE", "STATUT FINAL"], 1):
        cell(ws, r_b, c, h, f=fill(DARK_GREY), ft=font(WHITE, 10, bold=True),
             al=align("center"), bd=border(), locked=True)
    for i in range(1, 9):
        r = r_b + i; hrow(ws, r, 33)
        cell(ws, r, 1, f"Mission {i}", f=fill(LOCKED_BG), ft=font(bold=True,sz=10),
             al=align("center"), bd=border(), locked=True)
        cell(ws, r, 2, "", f=fill(INPUT_YLW), ft=font(sz=10),
             al=align("left","center",True), bd=border(), locked=False)
        cell(ws, r, 3, "☐ Non commencé", f=fill(INPUT_YLW), ft=font(sz=10),
             al=align("center"), bd=border(), locked=False)

    spacer(ws, r_b+9, 10)

    # Section C — Difficultés (SAISISSABLE)
    r_c = section_block(r_b+10, "SECTION C — Difficultés Rencontrées  (Facultatif)", ACCENT_BLUE)
    for r in range(r_c, r_c+3):
        hrow(ws, r, 33)
        merge(ws, r, 1, r, 3, "", f=fill(INPUT_YLW), ft=font(sz=10),
              al=align("left","center",True), locked=False)
        ws.cell(r, 1).border = border()

    spacer(ws, r_c+3, 10)

    # Section D — Besoins/Escalades (SAISISSABLE — fond orange pour alerte visuelle)
    r_d = section_block(r_c+4, "SECTION D — Besoins / Escalades  ⚠️ Déclenche une alerte si renseigné",
                         ORANGE_DARK)
    for r in range(r_d, r_d+3):
        hrow(ws, r, 33)
        merge(ws, r, 1, r, 3, "", f=fill(ORANGE_LIGHT), ft=font(sz=10),
              al=align("left","center",True), locked=False)
        ws.cell(r, 1).border = lborder(ORANGE_DARK)

    spacer(ws, r_d+3, 10)
    hrow(ws, r_d+4, 22)
    merge(ws, r_d+4, 1, r_d+4, 3,
          "⏰ RAPPEL : Clôture automatique à 18h00 — Verrouillage définitif de toutes les cellules",
          f=fill(RED_LIGHT), ft=font(RED_DARK, 9, italic=True), al=align("center"), locked=True)

    enable_protection(ws)

# ════════════════════════════════════════════════════════════════════════
# SHEET 3 — MES PERFORMANCES  (LECTURE SEULE INTÉGRALE)
# ════════════════════════════════════════════════════════════════════════
def sheet_performances(ws, dept):
    wcol(ws, "A", 28); wcol(ws, "B", 22); wcol(ws, "C", 38); wcol(ws, "D", 18)

    banner(ws, 1, 4, f"📊  MES PERFORMANCES — {dept.upper()}")
    hrow(ws, 2, 24)
    merge(ws, 2, 1, 2, 3, "RESPONSABLE : {{PRÉNOM NOM}}",
          f=fill(SUBHDR_BLUE), ft=font(WHITE, 11), al=align("left","center"), locked=True)
    cell(ws, 2, 4, "Mis à jour : {{TIMESTAMP}}",
         f=fill(SUBHDR_BLUE), ft=font(LIGHT_GOLD, 9, italic=True), al=align("right","center"), locked=True)
    spacer(ws, 3, 8)

    periods = [
        ("AUJOURD'HUI",  "{{TAUX_JOUR}}%",    "1565C0", "E3F2FD"),
        ("CETTE SEMAINE","{{TAUX_SEMAINE}}%",  "2E7D32", "E8F5E9"),
        ("CE MOIS",      "{{TAUX_MOIS}}%",     "F57F17", "FFF9C4"),
        ("CETTE ANNÉE",  "{{TAUX_ANNEE}}%",    "6A1B9A", "F3E5F5"),
    ]
    for i, (period, rate, fg, bg) in enumerate(periods):
        r = 4 + i*3
        hrow(ws, r, 28); hrow(ws, r+1, 18); hrow(ws, r+2, 8)
        cell(ws, r, 1, period, f=fill(fg), ft=font(WHITE, 11, bold=True),
             al=align("left","center"), bd=border(), locked=True)
        cell(ws, r, 2, rate, f=fill(fg), ft=font(WHITE, 15, bold=True),
             al=align("center","center"), bd=border(), locked=True)
        merge(ws, r, 3, r, 4, "████████████░░░░░░░░  {{BARRE}}",
              f=fill(bg), ft=font(sz=10), al=align("left","center"), locked=True)
        merge(ws, r+1, 1, r+1, 4, "Taux de complétion — Calculé automatiquement par BRASUS",
              f=fill(bg), ft=font(DARK_GREY, 8, italic=True), al=align("center"), locked=True)

    sep_r = 16; hrow(ws, sep_r, 12)
    merge(ws, sep_r, 1, sep_r, 4, "─"*90,
          f=fill(LIGHT_GREY), ft=font(MID_GREY, 8), al=align("center"), locked=True)

    section_header(ws, sep_r+1, 4, "OBJECTIF MENSUEL vs RÉALISÉ")
    for j, (label, val) in enumerate([
        ("MON OBJECTIF MENSUEL", "{{OBJECTIF_MENSUEL}}"),
        ("MON RÉSULTAT ACTUEL",  "{{RÉSULTAT_ACTUEL}}"),
        ("ÉCART",                "{{ÉCART}} ({{ÉCART_PCT}}%)"),
    ]):
        r = sep_r + 2 + j; hrow(ws, r, 28)
        cell(ws, r, 1, label, f=fill(LIGHT_GREY), ft=font(bold=True,sz=11),
             al=align("left","center"), bd=border(), locked=True)
        merge(ws, r, 2, r, 4, val, f=fill(LOCKED_BG), ft=font(italic=True,sz=11),
              al=align("center","center"), locked=True)
        ws.cell(r, 2).border = border()

    spacer(ws, sep_r+5, 12)
    section_header(ws, sep_r+6, 4, "MON PLAN D'AMÉLIORATION — Recommandations BRASUS", MID_BLUE)
    for k, (lbl, val) in enumerate([
        ("→ Recommandation 1", "{{RECO_1 — Générée par BRASUS}}"),
        ("→ Recommandation 2", "{{RECO_2 — Générée par BRASUS}}"),
        ("→ Recommandation 3", "{{RECO_3 — Générée par BRASUS}}"),
    ]):
        r = sep_r + 7 + k; hrow(ws, r, 36)
        cell(ws, r, 1, lbl, f=fill(LIGHT_BLUE), ft=font(MID_BLUE, 10, bold=True),
             al=align("left","center"), bd=border(), locked=True)
        # ⚠️ Recommandations générées par BRASUS → VERROUILLÉES (LOCKED_BG, pas INPUT_YLW)
        merge(ws, r, 2, r, 4, val, f=fill(LOCKED_BG), ft=font(DARK_GREY, 10, italic=True),
              al=align("left","center",True), locked=True)
        ws.cell(r, 2).border = border()

    footer_lock(ws, sep_r+11, 4)
    enable_protection(ws)

# ════════════════════════════════════════════════════════════════════════
# SHEET 4 — MON HISTORIQUE  (LECTURE SEULE INTÉGRALE)
# ════════════════════════════════════════════════════════════════════════
def sheet_historique(ws, dept):
    wcol(ws, "A", 14); wcol(ws, "B", 22); wcol(ws, "C", 24)
    wcol(ws, "D", 14); wcol(ws, "E", 36); wcol(ws, "F", 36)

    banner(ws, 1, 6, f"📅  MON HISTORIQUE — {dept.upper()}")
    hrow(ws, 2, 22)
    merge(ws, 2, 1, 2, 6,
          "🔒 Données en lecture seule — Archivage automatique par BRASUS — Aucune modification manuelle autorisée",
          f=fill(LOCKED_BG), ft=font(DARK_GREY, 9, italic=True), al=align("center","center"), locked=True)
    spacer(ws, 3, 10)

    hrow(ws, 4, 24)
    tabs = [("VUE JOURNALIÈRE (90 jours)", MID_BLUE, WHITE),
            ("VUE HEBDOMADAIRE (52 sem.)", LIGHT_GREY, DARK_GREY),
            ("VUE MENSUELLE (24 mois)",    LIGHT_GREY, DARK_GREY)]
    for t_i, (label, bg, fg) in enumerate(tabs):
        c = 1 + t_i*2
        merge(ws, 4, c, 4, c+1, f"  {label}", f=fill(bg), ft=font(fg, 10, bold=(t_i==0)),
              al=align("left","center"), locked=True)

    spacer(ws, 5, 8)
    hrow(ws, 6, 28)
    for c, h in enumerate(["DATE","MISSIONS ASSIGNÉES","MISSIONS RÉALISÉES","TAUX","POINTS FORTS","POINTS À AMÉLIORER"], 1):
        cell(ws, 6, c, h, f=fill(DARK_GREY), ft=font(WHITE, 10, bold=True),
             al=align("center","center",True), bd=border(), locked=True)

    today = date.today()
    for d in range(1, 16):
        r = 6 + d; hrow(ws, r, 28)
        day = today - timedelta(days=d)
        cell(ws, r, 1, day.strftime("%d/%m/%Y"),
             f=fill(LOCKED_BG), ft=font(sz=10), al=align("center"), bd=border(), locked=True)
        cell(ws, r, 2, "{{N — Généré BRASUS}}",
             f=fill(LOCKED_BG), ft=font(italic=True,sz=9), al=align("center"), bd=border(), locked=True)
        cell(ws, r, 3, "{{N — Généré BRASUS}}",
             f=fill(LOCKED_BG), ft=font(italic=True,sz=9), al=align("center"), bd=border(), locked=True)
        cell(ws, r, 4, "{{X%}}",
             f=fill(LOCKED_BG), ft=font(bold=True,sz=10), al=align("center"), bd=border(), locked=True)
        cell(ws, r, 5, "{{Points forts — Analysé par BRASUS}}",
             f=fill(LOCKED_BG), ft=font(italic=True,sz=9), al=align("left","center",True), bd=border(), locked=True)
        cell(ws, r, 6, "{{Axes d'amélioration — Analysé par BRASUS}}",
             f=fill(LOCKED_BG), ft=font(italic=True,sz=9), al=align("left","center",True), bd=border(), locked=True)

    hrow(ws, 22, 22)
    merge(ws, 22, 1, 22, 6,
          "Historique journalier : 90 jours  |  Hebdomadaire : 52 semaines  |  Mensuel : 24 mois",
          f=fill(LIGHT_BLUE), ft=font(MID_BLUE, 9, italic=True), al=align("center"), locked=True)

    footer_lock(ws, 23, 6)
    enable_protection(ws)

# ════════════════════════════════════════════════════════════════════════
# SHEET 5 — PLAN D'AMÉLIORATION  (LECTURE SEULE INTÉGRALE)
# ════════════════════════════════════════════════════════════════════════
def sheet_plan(ws, dept):
    wcol(ws, "A", 8); wcol(ws, "B", 66); wcol(ws, "C", 22)

    banner(ws, 1, 3, f"🚀  PLAN D'AMÉLIORATION — {dept.upper()}")
    hrow(ws, 2, 24)
    merge(ws, 2, 1, 2, 2, "Semaine du {{DATE_DEBUT}} au {{DATE_FIN}}",
          f=fill(SUBHDR_BLUE), ft=font(WHITE, 11), al=align("left","center"), locked=True)
    cell(ws, 2, 3, "Préparé par BRASUS pour {{PRÉNOM}}",
         f=fill(SUBHDR_BLUE), ft=font(LIGHT_GOLD, 9, italic=True), al=align("right","center"), locked=True)
    spacer(ws, 3, 12)

    section_header(ws, 4, 3, "DIAGNOSTIC")
    for j, (label, val) in enumerate([
        ("Taux d'exécution cette semaine :", "{{TAUX_EXECUTION_SEMAINE}}%"),
        ("Tendance :", "{{↑ En progression / → Stable / ↓ En baisse}}"),
        ("Analyse BRASUS :", "{{DIAGNOSTIC_CLAUDE — Généré automatiquement}}"),
    ]):
        r = 5+j; hrow(ws, r, 30)
        cell(ws, r, 1, label, f=fill(LIGHT_GREY), ft=font(bold=True,sz=10),
             al=align("left","center"), bd=border(), locked=True)
        merge(ws, r, 2, r, 3, val, f=fill(LOCKED_BG), ft=font(italic=True,sz=10),
              al=align("left","center",True), locked=True)
        ws.cell(r, 2).border = border()

    spacer(ws, 8, 12)
    section_header(ws, 9, 3, "VOS 3 AXES D'AMÉLIORATION PRIORITAIRES", MID_BLUE)
    for k in range(1, 4):
        r = 9+k; hrow(ws, r, 45)
        cell(ws, r, 1, str(k), f=fill(MID_BLUE), ft=font(WHITE, 14, bold=True),
             al=align("center","center"), bd=border(), locked=True)
        merge(ws, r, 2, r, 3, f"{{{{AXE_{k} — Concret et actionnable — Généré par BRASUS}}}}",
              f=fill(LOCKED_BG), ft=font(DARK_GREY, 10, italic=True),
              al=align("left","center",True), locked=True)
        ws.cell(r, 2).border = border()

    spacer(ws, 13, 12)
    section_header(ws, 14, 3, "OBJECTIF DE LA SEMAINE PROCHAINE", GREEN_MID)
    hrow(ws, 15, 52)
    merge(ws, 15, 1, 15, 3, "{{OBJECTIF_SPÉCIFIQUE_MESURABLE — Défini par BRASUS}}",
          f=fill(LOCKED_BG), ft=font(DARK_GREY, 11, italic=True),
          al=align("left","center",True), locked=True)
    ws.cell(15, 1).border = border()

    spacer(ws, 16, 12)
    section_header(ws, 17, 3, "RESSOURCES RECOMMANDÉES", PURPLE_MID)
    for r_i in range(1, 4):
        r = 17+r_i; hrow(ws, r, 36)
        merge(ws, r, 1, r, 3,
              f"▸  {{{{RESSOURCE_{r_i} — Template / guide / action recommandée par BRASUS}}}}",
              f=fill(PURPLE_LIGHT), ft=font(DARK_GREY, 10, italic=True),
              al=align("left","center",True), locked=True)
        ws.cell(r, 1).border = lborder(PURPLE_MID)

    spacer(ws, 21, 12)
    footer_lock(ws, 22, 3, "🔒 Plan généré automatiquement chaque lundi à 07:00 — Lecture seule intégrale")
    enable_protection(ws)

# ════════════════════════════════════════════════════════════════════════
# CRÉATION FICHIER DÉPARTEMENTAL
# ════════════════════════════════════════════════════════════════════════
DEPT_COLORS = {
    "COMMERCIAL":          "1565C0",
    "MARKETING":           "2E7D32",
    "RESSOURCES_HUMAINES": "6A1B9A",
    "FINANCE":             "C62828",
    "OPERATIONS":          "E65100",
}

def create_dept_file(dept_id, dept_label, output_dir):
    wb = openpyxl.Workbook()
    ws_default = wb.active

    ws1 = wb.create_sheet("1 - MISSIONS DU JOUR");       sheet_missions(ws1, dept_label)
    ws2 = wb.create_sheet("2 - EXÉCUTION JOURNALIÈRE");  sheet_execution(ws2, dept_label)
    ws3 = wb.create_sheet("3 - MES PERFORMANCES");       sheet_performances(ws3, dept_label)
    ws4 = wb.create_sheet("4 - MON HISTORIQUE");         sheet_historique(ws4, dept_label)
    ws5 = wb.create_sheet("5 - PLAN D'AMÉLIORATION");    sheet_plan(ws5, dept_label)

    wb.remove(ws_default)

    tc = DEPT_COLORS.get(dept_id, HDR_BLUE)
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = tc

    fname = f"BRASUS_{dept_id}.xlsx"
    wb.save(os.path.join(output_dir, fname))
    print(f"  ✅  {fname}")
    return fname

# ════════════════════════════════════════════════════════════════════════
# COCKPIT DG — toutes les feuilles en lecture seule intégrale
# ════════════════════════════════════════════════════════════════════════
def cockpit_dashboard(ws):
    ws.sheet_properties.tabColor = HDR_BLUE
    wcol(ws, "A", 5)
    wcol(ws, "B", 22); wcol(ws, "C", 22); wcol(ws, "D", 22); wcol(ws, "E", 22); wcol(ws, "F", 22)

    hrow(ws, 1, 50)
    merge(ws, 1, 1, 1, 6, "🏢  BRASUS — COCKPIT DIRECTION GÉNÉRALE",
          f=fill(HDR_BLUE), ft=font(WHITE, 18, bold=True), al=align("center","center"), locked=True)
    hrow(ws, 2, 26)
    merge(ws, 2, 1, 2, 4, "{{NOM_ENTREPRISE}}  |  Date : {{DATE}}  |  Heure : {{HEURE}}",
          f=fill(SUBHDR_BLUE), ft=font(WHITE, 11), al=align("left","center"), locked=True)
    merge(ws, 2, 5, 2, 6, "Accès réservé : DIRECTEUR GÉNÉRAL",
          f=fill(SUBHDR_BLUE), ft=font(GOLD, 10, bold=True), al=align("right","center"), locked=True)
    spacer(ws, 3, 12)

    kpis = [
        ("CA CE MOIS",           "{{X FCFA}}", "vs obj: {{±X%}}",   "1565C0", LIGHT_BLUE),
        ("TRÉSORERIE",           "{{X FCFA}}", "Seuil: {{OK/⚠️}}",  GREEN_MID, GREEN_LIGHT),
        ("TAUX EXÉCUTION GLOBAL","{{X%}}",     "Tendance {{↑↓→}}",  PURPLE_MID, PURPLE_LIGHT),
        ("ALERTES ACTIVES",      "{{N}}",      "🔴 {{N}} / 🟡 {{N}}",RED_DARK, RED_LIGHT),
    ]
    for i, (lbl, val, sub, fg, bg) in enumerate(kpis):
        c = 2 + i
        hrow(ws, 4, 18); hrow(ws, 5, 42); hrow(ws, 6, 22); hrow(ws, 7, 10)
        cell(ws, 4, c, lbl, f=fill(fg), ft=font(WHITE, 9, bold=True),
             al=align("center","center"), bd=border(), locked=True)
        cell(ws, 5, c, val, f=fill(bg), ft=font(fg, 16, bold=True),
             al=align("center","center"), bd=border(), locked=True)
        cell(ws, 6, c, sub, f=fill(bg), ft=font(DARK_GREY, 9, italic=True),
             al=align("center"), bd=border(), locked=True)

    spacer(ws, 8, 12)
    section_header(ws, 9, 6, "SYNTHÈSE PERFORMANCE PAR DÉPARTEMENT")
    hrow(ws, 10, 25)
    for c, h in enumerate(["","DÉPARTEMENT","AUJOURD'HUI","CETTE SEMAINE","CE MOIS","STATUT"], 1):
        cell(ws, 10, c, h, f=fill(DARK_GREY), ft=font(WHITE, 10, bold=True),
             al=align("center"), bd=border(), locked=True)

    depts_info = [
        ("💼","COMMERCIAL",         "E3F2FD"),
        ("📢","MARKETING",          "E8F5E9"),
        ("👥","RESSOURCES HUMAINES", "FFF9C4"),
        ("💰","FINANCE",            RED_LIGHT),
        ("⚙️","OPÉRATIONS",         ORANGE_LIGHT),
    ]
    for d_i, (icon, d_name, d_bg) in enumerate(depts_info):
        r = 11 + d_i; hrow(ws, r, 30)
        cell(ws, r, 1, icon, f=fill(LIGHT_GREY), ft=font(sz=13),
             al=align("center","center"), bd=border(), locked=True)
        cell(ws, r, 2, d_name, f=fill(d_bg), ft=font(bold=True,sz=10),
             al=align("left","center"), bd=border(), locked=True)
        for c in range(3, 6):
            cell(ws, r, c, "{{Auto}}", f=fill(LOCKED_BG), ft=font(italic=True,sz=11),
                 al=align("center"), bd=border(), locked=True)
        cell(ws, r, 6, "{{🟢/🟡/🔴}}", f=fill(LOCKED_BG), ft=font(sz=12),
             al=align("center"), bd=border(), locked=True)

    hrow(ws, 17, 22)
    merge(ws, 17, 1, 17, 6, "🔒 Données consolidées par BRASUS — Lecture seule intégrale",
          f=fill(LOCKED_BG), ft=font(DARK_GREY, 9, italic=True), al=align("center"), locked=True)

    enable_protection(ws)


def cockpit_dept_view(ws, dept_label, dept_icon, dept_color):
    ws.sheet_properties.tabColor = dept_color
    wcol(ws, "A", 33); wcol(ws, "B", 18); wcol(ws, "C", 18)
    wcol(ws, "D", 18); wcol(ws, "E", 18); wcol(ws, "F", 16)

    hrow(ws, 1, 40)
    merge(ws, 1, 1, 1, 6, f"{dept_icon}  VUE CONSOLIDÉE — {dept_label.upper()}",
          f=fill(dept_color), ft=font(WHITE, 15, bold=True), al=align("center","center"), locked=True)
    hrow(ws, 2, 20)
    merge(ws, 2, 1, 2, 6,
          "🔒 Vue agrégée — Données brutes du département non accessibles via ce fichier",
          f=fill(LOCKED_BG), ft=font(DARK_GREY, 9, italic=True), al=align("center","center"), locked=True)
    spacer(ws, 3, 8)

    hrow(ws, 4, 28)
    for c, h in enumerate(["INDICATEUR","HIER","CETTE SEMAINE","CE MOIS","OBJECTIF","ÉCART"], 1):
        cell(ws, 4, c, h, f=fill(DARK_GREY), ft=font(WHITE, 10, bold=True),
             al=align("center","center",True), bd=border(), locked=True)

    kpi_names = [
        "KPI Principal 1", "KPI Principal 2", "KPI Principal 3",
        "Taux d'exécution missions", "Avancement objectif mensuel", "Alertes actives",
    ]
    for k_i, kpi in enumerate(kpi_names):
        r = 5 + k_i; hrow(ws, r, 28)
        cell(ws, r, 1, kpi, f=fill(LOCKED_BG), ft=font(bold=True,sz=10),
             al=align("left","center"), bd=border(), locked=True)
        for c in range(2, 7):
            cell(ws, r, c, "{{Auto}}", f=fill(LOCKED_BG), ft=font(italic=True,sz=10),
                 al=align("center"), bd=border(), locked=True)

    spacer(ws, 11, 12)
    hrow(ws, 12, 28)
    merge(ws, 12, 1, 12, 6, "  PERFORMANCE RESPONSABLE DE DÉPARTEMENT",
          f=fill(dept_color), ft=font(WHITE, 11, bold=True), al=align("left","center"), locked=True)

    for r_i, (lbl, val) in enumerate([
        ("Responsable",               "{{NOM_RESPONSABLE}}"),
        ("Taux exécution aujourd'hui","{{X%}}"),
        ("Taux exécution semaine",    "{{X%}}"),
        ("Taux exécution mois",       "{{X%}}"),
        ("Missions en retard",        "{{N}}"),
        ("Dernière saisie",           "{{TIMESTAMP}}"),
    ]):
        r = 13 + r_i; hrow(ws, r, 26)
        cell(ws, r, 1, lbl, f=fill(LIGHT_GREY), ft=font(bold=True,sz=10),
             al=align("left","center"), bd=border(), locked=True)
        merge(ws, r, 2, r, 6, val, f=fill(LOCKED_BG), ft=font(italic=True,sz=10),
              al=align("center","center"), locked=True)
        ws.cell(r, 2).border = border()

    enable_protection(ws)


def cockpit_rapports_ia(ws):
    ws.sheet_properties.tabColor = HDR_BLUE
    wcol(ws, "A", 8); wcol(ws, "B", 75)

    banner(ws, 1, 2, "🤖  ANALYSE BRASUS — RAPPORTS INTELLIGENCE ARTIFICIELLE")
    hrow(ws, 2, 22)
    merge(ws, 2, 1, 2, 2, "Analyse générée par BRASUS — Date : {{DATE}}  |  Heure : {{HEURE}}",
          f=fill(SUBHDR_BLUE), ft=font(WHITE, 10), al=align("left","center"), locked=True)
    spacer(ws, 3, 10)

    sections = [
        ("SITUATION GLOBALE",                    MID_BLUE,    LIGHT_BLUE,
         ["{{SYNTHÈSE_GLOBALE_CLAUDE — 3 lignes}}"]),
        ("POINTS FORTS DE LA PÉRIODE",           GREEN_MID,   GREEN_LIGHT,
         ["→  {{POINT_FORT_1}}", "→  {{POINT_FORT_2}}"]),
        ("POINTS DE VIGILANCE",                  ORANGE_DARK, ORANGE_LIGHT,
         ["⚠️  {{ALERTE_1}}  —  Criticité : {{NIVEAU}}", "⚠️  {{ALERTE_2}}  —  Criticité : {{NIVEAU}}"]),
        ("RECOMMANDATIONS STRATÉGIQUES",         PURPLE_MID,  PURPLE_LIGHT,
         ["1.  {{ACTION_1}}  |  Délai : {{D}}  |  Responsable : {{DEPT}}",
          "2.  {{ACTION_2}}  |  Délai : {{D}}  |  Responsable : {{DEPT}}",
          "3.  {{ACTION_3}}  |  Délai : {{D}}  |  Responsable : {{DEPT}}"]),
        ("INDICATEURS À SURVEILLER CETTE SEMAINE", RED_DARK,  RED_LIGHT,
         ["▸  {{KPI_1}}  →  Seuil : {{S1}}", "▸  {{KPI_2}}  →  Seuil : {{S2}}", "▸  {{KPI_3}}  →  Seuil : {{S3}}"]),
    ]

    r = 4
    for title, hdr_c, body_c, lines in sections:
        hrow(ws, r, 28)
        merge(ws, r, 1, r, 2, f"  {title}", f=fill(hdr_c), ft=font(WHITE, 11, bold=True),
              al=align("left","center"), locked=True)
        r += 1
        for line in lines:
            hrow(ws, r, 40)
            merge(ws, r, 1, r, 2, f"  {line}", f=fill(body_c), ft=font(italic=True,sz=10),
                  al=align("left","center",True), locked=True)
            ws.cell(r, 1).border = lborder(hdr_c)
            r += 1
        hrow(ws, r, 10); r += 1

    footer_lock(ws, r, 2, "🔒 Rapport généré automatiquement par BRASUS — Mis à jour quotidiennement à 18h30")
    enable_protection(ws)


def cockpit_plan_actions(ws):
    ws.sheet_properties.tabColor = HDR_BLUE
    wcol(ws, "A", 6); wcol(ws, "B", 46); wcol(ws, "C", 26)
    wcol(ws, "D", 14); wcol(ws, "E", 16); wcol(ws, "F", 18)

    banner(ws, 1, 6, "📋  PLAN D'ACTIONS PRIORITAIRES — DIRECTION GÉNÉRALE")
    hrow(ws, 2, 22)
    merge(ws, 2, 1, 2, 6, "Généré automatiquement chaque lundi à 07:00 par BRASUS",
          f=fill(SUBHDR_BLUE), ft=font(WHITE, 10, italic=True), al=align("left","center"), locked=True)
    spacer(ws, 3, 10)

    hrow(ws, 4, 28)
    for c, h in enumerate(["#","ACTION","DÉPARTEMENT CONCERNÉ","URGENCE","ÉCHÉANCE","STATUT"], 1):
        cell(ws, 4, c, h, f=fill(DARK_GREY), ft=font(WHITE, 10, bold=True),
             al=align("center","center",True), bd=border(), locked=True)

    urgences = [
        ("🔴 Critique", RED_LIGHT),
        ("🔴 Critique", RED_LIGHT),
        ("🟡 Haute",    LIGHT_GOLD),
        ("🟡 Haute",    LIGHT_GOLD),
        ("🟢 Standard", GREEN_LIGHT),
        ("🟢 Standard", GREEN_LIGHT),
    ]
    for i, (urg, ubg) in enumerate(urgences, 1):
        r = 4 + i; hrow(ws, r, 35)
        cell(ws, r, 1, str(i), f=fill(LOCKED_BG), ft=font(bold=True,sz=11),
             al=align("center","center"), bd=border(), locked=True)
        cell(ws, r, 2, f"{{{{ACTION_{i} — Générée par BRASUS}}}}",
             f=fill(LOCKED_BG), ft=font(italic=True,sz=10), al=align("left","center",True),
             bd=border(), locked=True)
        cell(ws, r, 3, "{{DÉPARTEMENT}}", f=fill(LOCKED_BG), ft=font(italic=True,sz=10),
             al=align("center"), bd=border(), locked=True)
        cell(ws, r, 4, urg, f=fill(ubg), ft=font(bold=True,sz=10),
             al=align("center"), bd=border(), locked=True)
        cell(ws, r, 5, "{{DATE}}", f=fill(LOCKED_BG), ft=font(italic=True,sz=10),
             al=align("center"), bd=border(), locked=True)
        # ── Seule la colonne STATUT est modifiable par le DG ──
        cell(ws, r, 6, "☐ À faire", f=fill(INPUT_YLW), ft=font(sz=10),
             al=align("center"), bd=border(), locked=False)

    hrow(ws, 11, 22)
    merge(ws, 11, 1, 11, 6,
          "Colonne STATUT (fond jaune) modifiable par le DG — Toutes les autres colonnes en lecture seule",
          f=fill(LIGHT_GREY), ft=font(DARK_GREY, 9, italic=True), al=align("center"), locked=True)

    enable_protection(ws)


def create_cockpit_file(output_dir):
    wb = openpyxl.Workbook()
    ws_default = wb.active

    ws1 = wb.create_sheet("1 - TABLEAU DE BORD");         cockpit_dashboard(ws1)
    ws2 = wb.create_sheet("2 - COMMERCIAL");               cockpit_dept_view(ws2, "Commercial",         "💼", "1565C0")
    ws3 = wb.create_sheet("3 - MARKETING");                cockpit_dept_view(ws3, "Marketing",           "📢", "2E7D32")
    ws4 = wb.create_sheet("4 - RESSOURCES HUMAINES");      cockpit_dept_view(ws4, "Ressources Humaines", "👥", "6A1B9A")
    ws5 = wb.create_sheet("5 - FINANCE");                  cockpit_dept_view(ws5, "Finance",             "💰", "C62828")
    ws6 = wb.create_sheet("6 - OPÉRATIONS");               cockpit_dept_view(ws6, "Opérations",          "⚙️", "E65100")
    ws7 = wb.create_sheet("7 - RAPPORTS IA");              cockpit_rapports_ia(ws7)
    ws8 = wb.create_sheet("8 - PLAN D'ACTIONS");           cockpit_plan_actions(ws8)
    wb.remove(ws_default)

    fname = "BRASUS_COCKPIT_DG.xlsx"
    wb.save(os.path.join(output_dir, fname))
    print(f"  ✅  {fname}")
    return fname

# ════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════
DEPARTMENTS = [
    ("COMMERCIAL",          "Commercial"),
    ("MARKETING",           "Marketing"),
    ("RESSOURCES_HUMAINES", "Ressources Humaines"),
    ("FINANCE",             "Finance"),
    ("OPERATIONS",          "Opérations"),
]

def main():
    out = "/home/user/BRAZUS-OS-2026-REVU/brasus/skills/google-workspace/sheets"
    os.makedirs(out, exist_ok=True)

    print("\n🚀  BRASUS v2.1 — Génération des fichiers Google Sheets (avec protection réelle)")
    print("    Digital Mastery Agency (DMA) — Abidjan, Côte d'Ivoire")
    print("─" * 62)
    print("Fichiers départementaux :")
    for dept_id, dept_label in DEPARTMENTS:
        create_dept_file(dept_id, dept_label, out)

    print("\nCockpit Direction Générale :")
    create_cockpit_file(out)

    files = sorted(os.listdir(out))
    print(f"\n✅  {len(files)} fichiers générés dans :\n   {out}\n")
    for f in files:
        size = os.path.getsize(os.path.join(out, f)) // 1024
        print(f"   📄  {f}  ({size} Ko)")

    print(f"""
╔══════════════════════════════════════════════════════════════╗
║  PROTECTION CELLULES — RÉSUMÉ                                ║
║  Fond JAUNE  (INPUT_YLW)  = SAISISSABLE par l'utilisateur   ║
║  Fond GRIS   (LOCKED_BG)  = VERROUILLÉ — lecture seule      ║
║  Mot de passe admin       = {ADMIN_PWD:<30s} ║
║  ⚠️  À changer avant déploiement client                      ║
╠══════════════════════════════════════════════════════════════╣
║  IMPORT DANS GOOGLE SHEETS                                   ║
║  1. drive.google.com → + Nouveau → Importer un fichier       ║
║  2. Choisir "Remplacer la feuille de calcul"                 ║
║  3. Google Sheets respecte les plages protégées Excel        ║
║  4. Ajuster les permissions via Données → Protéger les feu.  ║
╚══════════════════════════════════════════════════════════════╝
""")

if __name__ == "__main__":
    main()
